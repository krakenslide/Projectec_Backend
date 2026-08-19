from datetime import datetime, timedelta, time, timezone, date
from uuid import UUID

from fastapi import status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.activity import Activity
from app.models.comment import Comment
from app.models.project import Project
from app.models.ticket import Ticket
from app.models.user import User
from app.models.user_organization import UserOrganization

from .schemas import (
    DeveloperCommentSummary,
    DeveloperDailySummary,
    DeveloperDailySummaryResponse,
    DeveloperTicketSummary,
)


from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from sqlalchemy import select

from app.models.role import Role
from app.models.user_project import UserProject

def get_report_window() -> tuple[datetime, datetime]:
    """
    Report runs from:
        Previous day 09:00
        ->
        Current day 09:00

    If the current time is before 09:00, the window is:
        Previous day 09:00
        ->
        Today 09:00
    """

    now = datetime.now(timezone.utc)

    today_9am = datetime.combine(
        now.date(),
        time(7, 0),
        tzinfo=timezone.utc,
    )

    if now < today_9am:
        end = today_9am
        start = end - timedelta(days=1)
    else:
        start = today_9am
        end = today_9am + timedelta(days=1)

    return start, end


# def get_report_window() -> tuple[datetime, datetime]:
#     now = datetime.now(timezone.utc)

#     today_9am = datetime.combine(
#         now.date(),
#         time(9, 0),
#         tzinfo=timezone.utc,
#     )

#     if now < today_9am:
#         end = today_9am
#         start = end - timedelta(days=1)
#     else:
#         start = today_9am - timedelta(days=1)
#         end = today_9am

#     return start, end


async def get_developer_daily_summary(
    db: AsyncSession,
    organization_id: UUID,
    user_id: UUID,
    current_user: User,
) -> DeveloperDailySummaryResponse:

    developer_membership = await db.scalar(
        select(UserOrganization).where(
            UserOrganization.organization_id == organization_id,
            UserOrganization.user_id == user_id,
        )
    )

    if developer_membership is None:
        raise ValueError(
            "User is not a member of this organization."
        )

    developer = await db.scalar(
        select(User).where(
            User.id == user_id,
        )
    )

    if developer is None:
        raise ValueError("User not found.")

    start, end = get_report_window()

    # ---------------------------------------------------------
    # Get all tickets assigned to this developer
    # within this organization.
    # ---------------------------------------------------------

    # result = await db.execute(
    #     select(
    #         Ticket,
    #         Project.name,
    #     )
    #     .join(
    #         Project,
    #         Project.id == Ticket.project_id,
    #     )
    #     .where(
    #         Project.organization_id == organization_id,
    #         Ticket.assigned_to == user_id,
    #     )
    # )

    result = await db.execute(
            select(
                Ticket,
                Project.name,
            )
            .join(
                Project,
                Project.id == Ticket.project_id,
            )
            .where(
                Project.organization_id == organization_id
            )
        )

    ticket_rows = result.all()
    print("WHYNOT")
    # print(ticket_rows[0].__dict__)
    for row in ticket_rows:
        print("WHYNOT2")
        print(row.Ticket.__dict__)
        #print(row.ticket_id)

    report: list[DeveloperTicketSummary] = []

    for ticket, project_name in ticket_rows:
        print("SHERLOCK")
        print(ticket.id)
        print(start)
        print(end)
        activities = (
            await db.scalars(
                select(Activity)
                .where(
                    Activity.ticket_id == ticket.id,
                    Activity.created_at >= start,
                    Activity.created_at < end,
                    Activity.created_by == user_id,
                )
                .order_by(Activity.created_at.asc())
            )
        ).all()

        comments = (
            await db.scalars(
                select(Comment)
                .where(
                    Comment.ticket_id == ticket.id,
                    Comment.created_by == user_id,
                    Comment.created_at >= start,
                    Comment.created_at < end,
                    Comment.is_deleted.is_(False),
                )
                .order_by(Comment.created_at.asc())
            )
        ).all()

        status_changed = False
        finished = False
        hours_logged = 0
        previous_status = ""


        # 3 conditions for status to be changed either 
        # 1. hours were logged
        # 2. ticket status was changed open to close etc
        # 3. Comment was done
        print("WHAT THE HELL")
        print(len(activities))
        for activity in activities:
            print("CHECK MAADY")
            print(activity.field_name)
            if activity.field_name == "status":

                status_changed = True
                previous_status = activity.old_value

                if (
                    activity.new_value is not None
                    and str(activity.new_value).lower() == "done"
                ):
                    finished = True

            elif activity.field_name == "hours_logged":
                status_changed = True
                previous_status = ticket.status

                try:
                    hours_logged += (
                        float(activity.new_value)
                        - float(activity.old_value)
                    )
                    hours_logged = int(hours_logged)
                except (TypeError, ValueError):
                    import traceback
                    traceback.print_exc()
                    pass
            else:
                previous_status = ticket.status

        comment_data = [
            DeveloperCommentSummary(
                id=comment.id,
                description=comment.description,
                created_at=comment.created_at,
            )
            for comment in comments
        ]
        if(len(comment_data)>0):
            status_changed = True

        if(status_changed == True or ticket.assigned_to == user_id ):
            report.append(
                DeveloperTicketSummary(
                    project_name=project_name,
                    ticket_name=ticket.title,
                    status_changed=status_changed,
                    previous_status= previous_status,
                    current_status=ticket.status,
                    finished=finished,
                    hours_logged=hours_logged,
                    comments=comment_data,
                )
            )


    total_tickets = len(report)

    tickets_finished = sum(
        1
        for ticket in report
        if ticket.finished
    )

    total_hours_logged = sum(
        ticket.hours_logged
        for ticket in report
    )

    total_comments = sum(
        len(ticket.comments)
        for ticket in report
    )

    summary = DeveloperDailySummary(
        user_id=developer.id,
        developer_name=developer.name,
        organization_id=organization_id,
        report_start=start,
        report_end=end,
        tickets=report,
        total_tickets=total_tickets,
        tickets_finished=tickets_finished,
        total_hours_logged=total_hours_logged,
        total_comments=total_comments,
    )

    return DeveloperDailySummaryResponse(
        success=True,
        status_code=status.HTTP_200_OK,
        message="Developer daily summary retrieved successfully.",
        data=summary,
    )


async def generate_organization_daily_report(
    db: AsyncSession,
    organization_id: UUID,
    current_user: User,
) -> BytesIO:

    result = await db.execute(
        select(User)
        .join(
            UserProject,
            UserProject.user_id == User.id,
        )
        .join(
            Project,
            Project.id == UserProject.project_id,
        )
        .join(
            Role,
            Role.id == UserProject.role_id,
        )
        .where(
            Project.organization_id == organization_id,
            Role.name == "Engineer",
        )
        .distinct()
    )

    engineers = result.scalars().all()

    workbook = Workbook()

    # Remove default sheet
    default_sheet = workbook.active
    workbook.remove(default_sheet)


    header_fill = PatternFill(
        "solid",
        fgColor="D9E2F3",
    )

    left_fill = PatternFill(
        "solid",
        fgColor="D9E2F3",
    )

    green_fill = PatternFill(
        "solid",
        fgColor="92D050",
    )

    yellow_fill = PatternFill(
        "solid",
        fgColor="FFD966",
    )

    blue_fill = PatternFill(
        "solid",
        fgColor="9DC3E6",
    )

    orange_fill = PatternFill(
        "solid",
        fgColor="F4B183",
    )

    white_fill = PatternFill(
        "solid",
        fgColor="FFFFFF",
    )

    thin = Side(style="thin")

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin,
    )

    for engineer in engineers:

        summary_response = await get_developer_daily_summary(
            db=db,
            organization_id=organization_id,
            user_id=engineer.id,
            current_user=current_user,
        )

        summary = summary_response.data

        # Excel sheet names have a 31-character limit
        sheet_name = engineer.name[:31] or "Developer"

        # Avoid duplicate sheet names
        existing_names = set(workbook.sheetnames)

        if sheet_name in existing_names:

            counter = 2

            while f"{sheet_name[:28]}_{counter}" in existing_names:
                counter += 1

            sheet_name = f"{sheet_name[:28]}_{counter}"

        worksheet = workbook.create_sheet(
            title=sheet_name,
        )

        labels = [
            "Project Name",
            "Ticket Name",
            "Status Changed",
            "Current Status",
            "Finished",
            "Hours Logged",
            "Comments",
        ]

        for row_number, label in enumerate(labels, start=1):

            cell = worksheet.cell(
                row=row_number,
                column=1,
            )

            cell.value = label
            cell.fill = left_fill
            cell.font = Font(bold=True)
            cell.border = border

        for column_number, ticket in enumerate(
            summary.tickets,
            start=2,
        ):

            values = [
                ticket.project_name,
                ticket.ticket_name,
                "Yes" if ticket.status_changed else "No",
                ticket.current_status,
                "Yes" if ticket.finished else "No",
                ticket.hours_logged,
                "\n".join(
                    comment.description
                    for comment in ticket.comments
                ),
            ]

            for row_number, value in enumerate(
                values,
                start=1,
            ):

                cell = worksheet.cell(
                    row=row_number,
                    column=column_number,
                )

                cell.value = value
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

                # Header row
                if row_number == 1:

                    cell.fill = header_fill
                    cell.font = Font(bold=True)

                # Current status
                elif row_number == 4:

                    status_value = str(value).lower()

                    if status_value == "done":
                        cell.fill = green_fill

                    elif status_value == "in progress":
                        cell.fill = yellow_fill

                    elif status_value == "testing":
                        cell.fill = blue_fill

                    elif "review" in status_value:
                        cell.fill = orange_fill

                    else:
                        cell.fill = white_fill

                else:
                    cell.fill = white_fill


        worksheet.column_dimensions["A"].width = 22

        for column_number in range(
            2,
            len(summary.tickets) + 2,
        ):
            worksheet.column_dimensions[
                get_column_letter(column_number)
            ].width = 24



    if not engineers:

        worksheet = workbook.create_sheet(
            title="No Engineers",
        )

        worksheet["A1"] = (
            "No engineers found in this organization."
        )


    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return output






from datetime import date, datetime
from uuid import UUID

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.ticket import Ticket
from app.models.user import User
from app.models.user_organization import UserOrganization


def to_date(
    value: date | datetime | None,
) -> date | None:
    """
    Convert a datetime/date value to a date.

    Database fields are expected to be datetime values,
    while the analytics API exposes date-level granularity.
    """

    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    return value


def get_effective_start_date(
    expected_start_date: date | datetime | None,
    actual_start_date: date | datetime | None,
) -> date | None:
    """
    Effective Gantt start date:

        min(expected_start_date, actual_start_date)

    If only one value exists, use that value.
    """

    expected_start_date = to_date(expected_start_date)
    actual_start_date = to_date(actual_start_date)

    if expected_start_date and actual_start_date:
        return min(
            expected_start_date,
            actual_start_date,
        )

    return expected_start_date or actual_start_date


def get_effective_end_date(
    expected_end_date: date | datetime | None,
    actual_end_date: date | datetime | None,
) -> date | None:
    """
    Effective Gantt end date:

        max(expected_end_date, actual_end_date)

    If only one value exists, use that value.
    """

    expected_end_date = to_date(expected_end_date)
    actual_end_date = to_date(actual_end_date)

    if expected_end_date and actual_end_date:
        return max(
            expected_end_date,
            actual_end_date,
        )

    return expected_end_date or actual_end_date


async def get_organization_engagement_gantt(
    session: AsyncSession,
    organization_id: UUID,
) -> list[dict]:
    """
    Generate engagement Gantt data for every member
    of an organization.

    A member's engagement consists of tickets assigned
    to that member.

    Effective ticket start:
        min(expected_start_date, actual_start_date)

    Effective ticket end:
        max(expected_end_date, actual_end_date)

    Members without tickets are also returned with an
    empty tickets list.
    """

    stmt = (
        select(
            User.id.label("member_id"),
            User.name.label("first_name"),

            Ticket.id.label("ticket_id"),
            Ticket.ticket_number.label("ticket_key"),
            Ticket.title.label("title"),

            Ticket.expected_start_date.label(
                "expected_start_date"
            ),
            Ticket.actual_start_date.label(
                "actual_start_date"
            ),

            Ticket.expected_end_date.label(
                "expected_end_date"
            ),
            Ticket.actual_end_date.label(
                "actual_end_date"
            ),
        )
        .join(
            UserOrganization,
            UserOrganization.user_id == User.id,
        )
        .outerjoin(
            Ticket,
            Ticket.assigned_to == User.id,
        )
        .where(
            UserOrganization.organization_id == organization_id,
        )
        .order_by(
            User.name,
            Ticket.expected_start_date,
        )
    )

    result = await session.execute(stmt)

    rows = result.all()

    members: dict[UUID, dict] = {}

    for row in rows:

        # --------------------------------------------------
        # Create member entry
        # --------------------------------------------------

        if row.member_id not in members:

            first_name = row.first_name or ""


            member_name = (
                f"{first_name}"
            ).strip()

            members[row.member_id] = {
                "member_id": row.member_id,
                "member_name": member_name,
                "tickets": [],
            }

        # --------------------------------------------------
        # No ticket assigned to this member
        # --------------------------------------------------

        if row.ticket_id is None:
            continue

        # --------------------------------------------------
        # Calculate effective Gantt dates
        # --------------------------------------------------

        start_date = get_effective_start_date(
            expected_start_date=row.expected_start_date,
            actual_start_date=row.actual_start_date,
        )

        end_date = get_effective_end_date(
            expected_end_date=row.expected_end_date,
            actual_end_date=row.actual_end_date,
        )

        # --------------------------------------------------
        # Cannot render a Gantt bar without both dates
        # --------------------------------------------------

        if start_date is None or end_date is None:
            continue

        # --------------------------------------------------
        # Add ticket
        # --------------------------------------------------

        members[row.member_id]["tickets"].append(
            {
                "ticket_id": row.ticket_id,
                "ticket_key": row.ticket_key,
                "title": row.title,

                # Effective Gantt dates
                "start_date": start_date,
                "end_date": end_date,

                # Original dates
                "expected_start_date": to_date(
                    row.expected_start_date
                ),
                "actual_start_date": to_date(
                    row.actual_start_date
                ),

                "expected_end_date": to_date(
                    row.expected_end_date
                ),
                "actual_end_date": to_date(
                    row.actual_end_date
                ),
            }
        )

    return list(members.values())