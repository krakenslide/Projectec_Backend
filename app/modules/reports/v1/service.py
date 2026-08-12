from datetime import datetime, timedelta, time, timezone
from uuid import UUID

from fastapi import status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.activity import Activity
from app.models.comment import Comment
from app.models.project import Project
from app.models.ticket import Ticket
from app.models.user import User
from app.models.user_organization import UserOrganization

from .schemas import (
    DeveloperCommentSummary,
    DeveloperDailySummary,
    DeveloperDailySummaryResponse,
    DeveloperTicketSummary,
)


from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from sqlalchemy import select

from app.models.role import Role
from app.models.user_project import UserProject

def get_report_window() -> tuple[datetime, datetime]:
    """
    Returns the reporting window used for the daily developer report.

    Report runs from:
        Previous day 09:00
        ->
        Current day 09:00

    If the current time is before 09:00, the window is:
        Previous day 09:00
        ->
        Today 09:00
    """

    now = datetime.now(timezone.utc)

    today_9am = datetime.combine(
        now.date(),
        time(9, 0),
        tzinfo=timezone.utc,
    )

    if now < today_9am:
        end = today_9am
        start = end - timedelta(days=1)
    else:
        start = today_9am
        end = today_9am + timedelta(days=1)

    return start, end


# def get_report_window() -> tuple[datetime, datetime]:
#     now = datetime.now(timezone.utc)

#     today_9am = datetime.combine(
#         now.date(),
#         time(9, 0),
#         tzinfo=timezone.utc,
#     )

#     if now < today_9am:
#         end = today_9am
#         start = end - timedelta(days=1)
#     else:
#         start = today_9am - timedelta(days=1)
#         end = today_9am

#     return start, end


async def get_developer_daily_summary(
    db: AsyncSession,
    organization_id: UUID,
    user_id: UUID,
    current_user: User,
) -> DeveloperDailySummaryResponse:

    # ---------------------------------------------------------
    # Validate that the developer belongs to the organization
    # ---------------------------------------------------------

    developer_membership = await db.scalar(
        select(UserOrganization).where(
            UserOrganization.organization_id == organization_id,
            UserOrganization.user_id == user_id,
        )
    )

    if developer_membership is None:
        raise ValueError(
            "User is not a member of this organization."
        )

    developer = await db.scalar(
        select(User).where(
            User.id == user_id,
        )
    )

    if developer is None:
        raise ValueError("User not found.")

    # ---------------------------------------------------------
    # Reporting window
    # ---------------------------------------------------------

    start, end = get_report_window()

    # ---------------------------------------------------------
    # Get all tickets assigned to this developer
    # within this organization.
    # ---------------------------------------------------------

    result = await db.execute(
        select(
            Ticket,
            Project.name,
        )
        .join(
            Project,
            Project.id == Ticket.project_id,
        )
        .where(
            Project.organization_id == organization_id,
            Ticket.assigned_to == user_id,
        )
    )

    ticket_rows = result.all()

    report: list[DeveloperTicketSummary] = []

    # ---------------------------------------------------------
    # Process every ticket
    # ---------------------------------------------------------

    for ticket, project_name in ticket_rows:

        # -----------------------------------------------------
        # Activities for this ticket during report window
        # -----------------------------------------------------

        activities = (
            await db.scalars(
                select(Activity)
                .where(
                    Activity.ticket_id == ticket.id,
                    Activity.created_at >= start,
                    Activity.created_at < end,
                )
                .order_by(Activity.created_at.asc())
            )
        ).all()

        # -----------------------------------------------------
        # Comments made by this developer during report window
        # -----------------------------------------------------

        comments = (
            await db.scalars(
                select(Comment)
                .where(
                    Comment.ticket_id == ticket.id,
                    Comment.created_by == user_id,
                    Comment.created_at >= start,
                    Comment.created_at < end,
                    Comment.is_deleted.is_(False),
                )
                .order_by(Comment.created_at.asc())
            )
        ).all()

        status_changed = False
        finished = False
        hours_logged = 0

        # -----------------------------------------------------
        # Analyze activities
        # -----------------------------------------------------

        for activity in activities:
            if activity.field_name == "status":

                status_changed = True

                if (
                    activity.new_value is not None
                    and str(activity.new_value).lower() == "done"
                ):
                    finished = True

            elif activity.field_name == "hours_logged":

                try:
                    hours_logged += (
                        float(activity.new_value)
                        - float(activity.old_value)
                    )
                    hours_logged = int(hours_logged)
                except (TypeError, ValueError):
                    import traceback
                    traceback.print_exc()
                    pass

        # -----------------------------------------------------
        # Convert comments to response objects
        # -----------------------------------------------------

        comment_data = [
            DeveloperCommentSummary(
                id=comment.id,
                description=comment.description,
                created_at=comment.created_at,
            )
            for comment in comments
        ]

        report.append(
            DeveloperTicketSummary(
                project_name=project_name,
                ticket_name=ticket.title,
                status_changed=status_changed,
                current_status=ticket.status,
                finished=finished,
                hours_logged=hours_logged,
                comments=comment_data,
            )
        )


    total_tickets = len(report)

    tickets_finished = sum(
        1
        for ticket in report
        if ticket.finished
    )

    total_hours_logged = sum(
        ticket.hours_logged
        for ticket in report
    )

    total_comments = sum(
        len(ticket.comments)
        for ticket in report
    )

    summary = DeveloperDailySummary(
        user_id=developer.id,
        developer_name=developer.name,
        organization_id=organization_id,
        report_start=start,
        report_end=end,
        tickets=report,
        total_tickets=total_tickets,
        tickets_finished=tickets_finished,
        total_hours_logged=total_hours_logged,
        total_comments=total_comments,
    )

    return DeveloperDailySummaryResponse(
        success=True,
        status_code=status.HTTP_200_OK,
        message="Developer daily summary retrieved successfully.",
        data=summary,
    )


async def generate_organization_daily_report(
    db: AsyncSession,
    organization_id: UUID,
    current_user: User,
) -> BytesIO:

    # ---------------------------------------------------------
    # Get all engineers belonging to projects in this
    # organization.
    # ---------------------------------------------------------

    result = await db.execute(
        select(User)
        .join(
            UserProject,
            UserProject.user_id == User.id,
        )
        .join(
            Project,
            Project.id == UserProject.project_id,
        )
        .join(
            Role,
            Role.id == UserProject.role_id,
        )
        .where(
            Project.organization_id == organization_id,
            Role.name == "Engineer",
        )
        .distinct()
    )

    engineers = result.scalars().all()

    # ---------------------------------------------------------
    # Create workbook
    # ---------------------------------------------------------

    workbook = Workbook()

    # Remove default sheet
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    # ---------------------------------------------------------
    # Styles
    # ---------------------------------------------------------

    header_fill = PatternFill(
        "solid",
        fgColor="D9E2F3",
    )

    left_fill = PatternFill(
        "solid",
        fgColor="D9E2F3",
    )

    green_fill = PatternFill(
        "solid",
        fgColor="92D050",
    )

    yellow_fill = PatternFill(
        "solid",
        fgColor="FFD966",
    )

    blue_fill = PatternFill(
        "solid",
        fgColor="9DC3E6",
    )

    orange_fill = PatternFill(
        "solid",
        fgColor="F4B183",
    )

    white_fill = PatternFill(
        "solid",
        fgColor="FFFFFF",
    )

    thin = Side(style="thin")

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin,
    )

    # ---------------------------------------------------------
    # Generate sheet for every engineer
    # ---------------------------------------------------------

    for engineer in engineers:

        summary_response = await get_developer_daily_summary(
            db=db,
            organization_id=organization_id,
            user_id=engineer.id,
            current_user=current_user,
        )

        summary = summary_response.data

        # Excel sheet names have a 31-character limit
        sheet_name = engineer.name[:31] or "Developer"

        # Avoid duplicate sheet names
        existing_names = set(workbook.sheetnames)

        if sheet_name in existing_names:

            counter = 2

            while f"{sheet_name[:28]}_{counter}" in existing_names:
                counter += 1

            sheet_name = f"{sheet_name[:28]}_{counter}"

        worksheet = workbook.create_sheet(
            title=sheet_name,
        )

        # -----------------------------------------------------
        # Headers
        # -----------------------------------------------------

        labels = [
            "Project Name",
            "Ticket Name",
            "Status Changed",
            "Current Status",
            "Finished",
            "Hours Logged",
            "Comments",
        ]

        for row_number, label in enumerate(labels, start=1):

            cell = worksheet.cell(
                row=row_number,
                column=1,
            )

            cell.value = label
            cell.fill = left_fill
            cell.font = Font(bold=True)
            cell.border = border

        # -----------------------------------------------------
        # Ticket data
        # -----------------------------------------------------

        for column_number, ticket in enumerate(
            summary.tickets,
            start=2,
        ):

            values = [
                ticket.project_name,
                ticket.ticket_name,
                "Yes" if ticket.status_changed else "No",
                ticket.current_status,
                "Yes" if ticket.finished else "No",
                ticket.hours_logged,
                "\n".join(
                    comment.description
                    for comment in ticket.comments
                ),
            ]

            for row_number, value in enumerate(
                values,
                start=1,
            ):

                cell = worksheet.cell(
                    row=row_number,
                    column=column_number,
                )

                cell.value = value
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

                # Header row
                if row_number == 1:

                    cell.fill = header_fill
                    cell.font = Font(bold=True)

                # Current status
                elif row_number == 4:

                    status_value = str(value).lower()

                    if status_value == "done":
                        cell.fill = green_fill

                    elif status_value == "in progress":
                        cell.fill = yellow_fill

                    elif status_value == "testing":
                        cell.fill = blue_fill

                    elif "review" in status_value:
                        cell.fill = orange_fill

                    else:
                        cell.fill = white_fill

                else:
                    cell.fill = white_fill

        # -----------------------------------------------------
        # Column widths
        # -----------------------------------------------------

        worksheet.column_dimensions["A"].width = 22

        for column_number in range(
            2,
            len(summary.tickets) + 2,
        ):
            worksheet.column_dimensions[
                get_column_letter(column_number)
            ].width = 24

    # ---------------------------------------------------------
    # If organization has no engineers
    # ---------------------------------------------------------

    if not engineers:

        worksheet = workbook.create_sheet(
            title="No Engineers",
        )

        worksheet["A1"] = (
            "No engineers found in this organization."
        )

    # ---------------------------------------------------------
    # Write workbook to memory
    # ---------------------------------------------------------

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return output