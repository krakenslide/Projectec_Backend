# from collections import defaultdict
# from datetime import datetime, timedelta, time

# from openpyxl import Workbook
# from openpyxl.styles import Font

# from sqlalchemy import and_

# from app.core.database import SessionLocal

# from app.models.user import User
# from app.models.role import Role
# from app.models.project import Project
# from app.models.ticket import Ticket
# from app.models.comment import Comment
# from app.models.activity import Activity
# from app.models.user_project import UserProject


# ORGANIZATION_ID = "030a820b-4a61-4f73-917c-0c1c42568df9"


# def get_report_window():
#     now = datetime.now()

#     today_9 = datetime.combine(now.date(), time(9, 0))

#     if now < today_9:
#         end_time = today_9
#         start_time = end_time - timedelta(days=1)
#     else:
#         start_time = today_9 - timedelta(days=1)
#         end_time = today_9

#     return start_time, end_time


# def generate_report():

#     db = SessionLocal()

#     start_time, end_time = get_report_window()

#     engineers = (
#         db.query(
#             User.id,
#             User.name,
#             User.email,
#         )
#         .join(UserProject, User.id == UserProject.user_id)
#         .join(Role, Role.id == UserProject.role_id)
#         .join(Project, Project.id == UserProject.project_id)
#         .filter(
#             Project.organization_id == ORGANIZATION_ID,
#             Role.name == "Engineer",
#         )
#         .distinct()
#         .all()
#     )

#     workbook = Workbook()

#     workbook.remove(workbook.active)

#     for engineer in engineers:

#         sheet = workbook.create_sheet(
#             title=engineer.name[:31]
#         )

#         headers = [
#             "Project Name",
#             "Ticket Name",
#             "Status Changed",
#             "Current Status",
#             "Finished",
#             "Hours Logged",
#             "Comments",
#         ]

#         for col, header in enumerate(headers, start=1):
#             cell = sheet.cell(row=1, column=col)
#             cell.value = header
#             cell.font = Font(bold=True)

#         tickets = (
#             db.query(
#                 Ticket,
#                 Project.name.label("project_name"),
#             )
#             .join(Project, Project.id == Ticket.project_id)
#             .filter(
#                 Ticket.assigned_to == engineer.id,
#                 Project.organization_id == ORGANIZATION_ID,
#             )
#             .all()
#         )

#         row_no = 2

#         for ticket, project_name in tickets:

#             activities = (
#                 db.query(Activity)
#                 .filter(
#                     Activity.ticket_id == ticket.id,
#                     Activity.created_at >= start_time,
#                     Activity.created_at < end_time,
#                 )
#                 .all()
#             )

#             comments = (
#                 db.query(Comment.description)
#                 .filter(
#                     Comment.ticket_id == ticket.id,
#                     Comment.created_by == engineer.id,
#                     Comment.created_at >= start_time,
#                     Comment.created_at < end_time,
#                 )
#                 .all()
#             )

#             status_changed = "No"
#             finished = "No"
#             hours_logged = 0

#             for activity in activities:

#                 if activity.field_name == "status":
#                     status_changed = "Yes"

#                     if (
#                         activity.new_value.lower()
#                         in [
#                             "done",
#                             "completed",
#                             "closed",
#                         ]
#                     ):
#                         finished = "Yes"

#                 elif activity.field_name == "hours_logged":

#                     try:
#                         hours_logged += (
#                             int(activity.new_value)
#                             - int(activity.old_value)
#                         )
#                     except:
#                         pass

#             comment_text = "\n".join(
#                 c.description if hasattr(c, "description") else c[0]
#                 for c in comments
#             )

#             sheet.cell(row=row_no, column=1).value = project_name
#             sheet.cell(row=row_no, column=2).value = ticket.title
#             sheet.cell(row=row_no, column=3).value = status_changed
#             sheet.cell(row=row_no, column=4).value = ticket.status
#             sheet.cell(row=row_no, column=5).value = finished
#             sheet.cell(row=row_no, column=6).value = hours_logged
#             sheet.cell(row=row_no, column=7).value = comment_text

#             row_no += 1

#         for column_cells in sheet.columns:

#             length = max(
#                 len(str(cell.value or ""))
#                 for cell in column_cells
#             )

#             sheet.column_dimensions[
#                 column_cells[0].column_letter
#             ].width = min(length + 5, 60)

#     filename = (
#         f"Daily_Report_"
#         f"{datetime.now().strftime('%Y%m%d')}.xlsx"
#     )

#     workbook.save(filename)

#     db.close()

#     print(f"Generated {filename}")


# if __name__ == "__main__":
#     generate_report()


"""
daily_engineer_report.py

Skeleton report generator for your ticketing system.

TODO:
- Replace ORGANIZATION_ID
- Verify model import paths if different.

This script:
- Calculates 9AM yesterday -> 9AM today
- Fetches engineers
- Builds one worksheet per engineer
- Formats worksheet like requested
"""

from collections import defaultdict
from datetime import datetime, timedelta, time
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import and_

from app.core.database import SessionLocal

from app.models.user import User
from app.models.role import Role
from app.models.project import Project
from app.models.ticket import Ticket
from app.models.comment import Comment
from app.models.activity import Activity
from app.models.user_project import UserProject

ORGANIZATION_ID = "030a820b-4a61-4f73-917c-0c1c42568df9"

HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
LEFT_FILL = PatternFill("solid", fgColor="D9E2F3")
GREEN_FILL = PatternFill("solid", fgColor="92D050")
YELLOW_FILL = PatternFill("solid", fgColor="FFD966")
BLUE_FILL = PatternFill("solid", fgColor="9DC3E6")
ORANGE_FILL = PatternFill("solid", fgColor="F4B183")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
THIN = Side(style="thin")
BORDER = Border(left=THIN,right=THIN,top=THIN,bottom=THIN)

def report_window():
    now = datetime.now()
    today9 = datetime.combine(now.date(), time(9,0))
    if now < today9:
        end = today9
        start = end - timedelta(days=1)
    else:
        start = today9 - timedelta(days=1)
        end = today9
    return start,end

def style_sheet(ws, rows):
    labels=[
        "Project Name","Ticket Name","Status Changed",
        "Current Status","Finished","Hours Logged","Comments"
    ]
    for r,label in enumerate(labels,1):
        c=ws.cell(row=r,column=1)
        c.value=label
        c.fill=LEFT_FILL
        c.font=Font(bold=True)
        c.border=BORDER

    for col,data in enumerate(rows,2):
        vals=[
            data["Project Name"],
            data["Ticket Name"],
            data["Status Changed"],
            data["Current Status"],
            data["Finished"],
            data["Hours Logged"],
            data["Comments"],
        ]
        for r,val in enumerate(vals,1):
            c=ws.cell(row=r,column=col)
            c.value=val
            c.border=BORDER
            c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
            if r==1:
                c.fill=HEADER_FILL
                c.font=Font(bold=True)
            elif r==4:
                s=str(val).lower()
                if s=="done":
                    c.fill=GREEN_FILL
                elif s=="in progress":
                    c.fill=YELLOW_FILL
                elif s=="testing":
                    c.fill=BLUE_FILL
                elif "review" in s:
                    c.fill=ORANGE_FILL
                else:
                    c.fill=WHITE_FILL
            else:
                c.fill=WHITE_FILL

    ws.column_dimensions["A"].width=22
    for c in range(2,len(rows)+2):
        ws.column_dimensions[get_column_letter(c)].width=24

def main():
    db=SessionLocal()
    start,end=report_window()

    engineers=(
        db.query(User)
        .join(UserProject,User.id==UserProject.user_id)
        .join(Role,Role.id==UserProject.role_id)
        .join(Project,Project.id==UserProject.project_id)
        .filter(Project.organization_id==ORGANIZATION_ID,
                Role.name=="Engineer")
        .distinct()
        .all()
    )

    wb=Workbook()
    wb.remove(wb.active)

    for eng in engineers:
        ws=wb.create_sheet(eng.name[:31])

        tickets=(
            db.query(Ticket,Project.name)
            .join(Project,Project.id==Ticket.project_id)
            .filter(Project.organization_id==ORGANIZATION_ID,
                    Ticket.assigned_to==eng.id)
            .all()
        )

        report=[]

        for ticket,project_name in tickets:
            acts=db.query(Activity).filter(
                Activity.ticket_id==ticket.id,
                Activity.created_at>=start,
                Activity.created_at<end
            ).all()

            comms=db.query(Comment).filter(
                Comment.ticket_id==ticket.id,
                Comment.created_by==eng.id,
                Comment.created_at>=start,
                Comment.created_at<end
            ).all()

            status_changed="No"
            finished="No"
            hours=0

            for a in acts:
                if a.field_name=="status":
                    status_changed="Yes"
                    if str(a.new_value).lower()=="done":
                        finished="Yes"
                elif a.field_name=="hours_logged":
                    try:
                        hours+=int(a.new_value)-int(a.old_value)
                    except:
                        pass

            report.append({
                "Project Name":project_name,
                "Ticket Name":ticket.title,
                "Status Changed":status_changed,
                "Current Status":ticket.status,
                "Finished":finished,
                "Hours Logged":hours,
                "Comments":"\n".join(c.description for c in comms)
            })

        style_sheet(ws,report)

    wb.save("Daily Engineer Report.xlsx")
    db.close()
    print("Done.")

if __name__=="__main__":
    main()